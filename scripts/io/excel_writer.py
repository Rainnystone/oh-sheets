import argparse
import sys
import json
import re
import openpyxl
from openpyxl.utils.cell import get_column_letter


CELL_REF_RE = re.compile(r"^[A-Za-z]+[0-9]+$")


def _is_cell_ref(value: str) -> bool:
    return bool(isinstance(value, str) and CELL_REF_RE.match(value))


def _shift_cell(cell_ref, row_offset=0, col_offset=0):
    if not _is_cell_ref(cell_ref):
        return None
    if row_offset == 0 and col_offset == 0:
        return cell_ref.upper()
    open_col = 0
    row_part = ""
    for ch in cell_ref:
        if ch.isdigit():
            row_part += ch
        else:
            open_col = open_col * 26 + (ord(ch.upper()) - ord("A") + 1)
    row_number = int(row_part or "1")
    new_col = open_col + col_offset
    new_row = row_number + row_offset
    if new_col <= 0 or new_row <= 0:
        return None
    return f"{get_column_letter(new_col)}{new_row}"


def _coerce_schema_bindings(schema_obj):
    bindings = []
    absolute_fields = {}
    pending_relative = []

    if not isinstance(schema_obj, dict):
        return bindings

    # New schema format: {"fields": {"field_name": {"cell": "B2", ...}}}
    fields = schema_obj.get("fields")
    if isinstance(fields, dict):
        for field_key, spec in fields.items():
            if not isinstance(spec, dict):
                continue
            cell_ref = spec.get("cell") or spec.get("address")
            if _is_cell_ref(cell_ref):
                absolute_fields[field_key] = cell_ref.upper()
                bindings.append((field_key, cell_ref.upper(), spec))
                continue
            if spec.get("relative_to"):
                pending_relative.append((field_key, spec))

        unresolved = True
        while unresolved and pending_relative:
            unresolved = False
            next_pending = []
            for field_key, spec in pending_relative:
                anchor = spec.get("relative_to")
                if anchor in absolute_fields:
                    anchor_cell = absolute_fields[anchor]
                    cell_ref = _shift_cell(
                        anchor_cell,
                        int(spec.get("row_offset", 0) or 0),
                        int(spec.get("col_offset", 0) or 0),
                    )
                    if cell_ref is None:
                        continue
                    absolute_fields[field_key] = cell_ref
                    bindings.append((field_key, cell_ref, spec))
                    unresolved = True
                else:
                    next_pending.append((field_key, spec))
            pending_relative = next_pending
        return bindings

    # Legacy format: {"B2": {"name": "Field_A", ...}}
    for cell_ref, spec in schema_obj.items():
        if cell_ref in ("fields", "meta", "version"):
            continue
        if _is_cell_ref(cell_ref):
            bindings.append((cell_ref, cell_ref.upper(), spec))
    return bindings


def write_excel(template_path, data_path, schema_path, output_path):
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        with open(data_path, 'r') as f:
            data = json.load(f)
            
        with open(schema_path, 'r') as f:
            schema = json.load(f)

        bindings = _coerce_schema_bindings(schema)
        for field_key, target_cell, spec in bindings:
            if field_key in data:
                ws[target_cell] = data[field_key]
                continue
            if isinstance(spec, dict):
                name = spec.get("name")
                if isinstance(name, str) and name in data:
                    ws[target_cell] = data[name]
                
        wb.save(output_path)
        sys.exit(0)
    except Exception as e:
        print(f"Error writing excel: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('--template', required=True)
    parser.add_argument('--data', required=True)
    parser.add_argument('--schema', required=True)
    parser.add_argument('--output', required=True)
    args = parser.parse_args()
    write_excel(args.template, args.data, args.schema, args.output)
