import argparse
import hashlib
import json
from pathlib import Path
import openpyxl


def _normalize_value(value):
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return str(value)
    return str(value).strip()


def _cell_neighbors(ws, row, column):
    max_row = ws.max_row
    max_column = ws.max_column
    return {
        "up": ws.cell(row - 1, column).coordinate if row > 1 else None,
        "down": ws.cell(row + 1, column).coordinate if row < max_row else None,
        "left": ws.cell(row, column - 1).coordinate if column > 1 else None,
        "right": ws.cell(row, column + 1).coordinate if column < max_column else None,
    }


def _collect_cells(ws):
    cells = []
    for row in range(1, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=column).value
            if value is None:
                continue
            text = _normalize_value(value)
            if text == "":
                continue
            coord = ws.cell(row=row, column=column).coordinate
            cells.append({
                "sheet": ws.title,
                "cell": coord,
                "row": row,
                "column": column,
                "value": text,
                "neighbors": _cell_neighbors(ws, row, column),
            })
    return cells


def _layout_signature_for_sheet(ws):
    occupied = []
    row_list = []
    for row in range(1, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=column).value
            if value is None:
                continue
            text = _normalize_value(value)
            if text == "":
                continue
            row_list.append((row, column, text))

    if not row_list:
        return {
            "name": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "used_rows": [],
            "used_cols": [],
            "occupied": [],
            "occupied_count": 0,
        }

    rows = [item[0] for item in row_list]
    cols = [item[1] for item in row_list]
    min_row = min(rows)
    min_col = min(cols)
    occupied = sorted((r - min_row, c - min_col) for r, c, _ in row_list)
    return {
        "name": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "used_rows": [min(rows), max(rows)],
        "used_cols": [min(cols), max(cols)],
        "occupied": occupied,
        "occupied_count": len(occupied),
    }


def _build_signature(payload):
    packed = json.dumps(payload, sort_keys=True, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(packed).hexdigest()


def _safe_set(value):
    if isinstance(value, list):
        return set(tuple(item) if isinstance(item, (list, tuple)) else item for item in value)
    return set()


def _jaccard(left, right):
    if not left and not right:
        return 1.0
    if not left or not right:
        return 0.0
    intersection = len(left & right)
    union = len(left | right)
    return intersection / union if union else 0.0


def _compare_sheet_sigs(base_sig, cand_sig):
    base_cells = _safe_set(base_sig.get("occupied", []))
    cand_cells = _safe_set(cand_sig.get("occupied", []))
    occupancy_jaccard = _jaccard(base_cells, cand_cells)

    base_rows = base_sig.get("used_rows") or [0, 0]
    cand_rows = cand_sig.get("used_rows") or [0, 0]
    base_cols = base_sig.get("used_cols") or [0, 0]
    cand_cols = cand_sig.get("used_cols") or [0, 0]

    base_h = max(base_rows[1] - base_rows[0], 0)
    cand_h = max(cand_rows[1] - cand_rows[0], 0)
    base_w = max(base_cols[1] - base_cols[0], 0)
    cand_w = max(cand_cols[1] - cand_cols[0], 0)

    row_delta = abs(base_h - cand_h) / (max(base_h, cand_h) or 1)
    col_delta = abs(base_w - cand_w) / (max(base_w, cand_w) or 1)
    shape_score = max(0.0, 1.0 - (row_delta + col_delta) / 2.0)
    score = occupancy_jaccard * 0.8 + shape_score * 0.2

    return {
        "name": base_sig.get("name"),
        "occupancy_jaccard": occupancy_jaccard,
        "shape_score": shape_score,
        "score": score,
        "name_match": base_sig.get("name") == cand_sig.get("name"),
    }


def _collect_sheet_map(profile):
    items = {}
    for sheet in profile.get("sheets", []):
        if isinstance(sheet, dict):
            sheet_name = sheet.get("name")
            if sheet_name:
                items[sheet_name] = sheet
    return items


def build_template_profile(template_path, max_cells=5000):
    wb = openpyxl.load_workbook(template_path, data_only=False)
    sheets = []
    all_cells = []
    for ws in wb.worksheets:
        sheet_cells = _collect_cells(ws)
        sheet_layout = _layout_signature_for_sheet(ws)
        all_cells.extend(sheet_cells)
        sheets.append({
            "name": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "layout_signature": sheet_layout,
        })

    profile = {
        "schema_version": "2.0",
        "template": str(Path(template_path).name),
        "signature_algorithm": "sha256",
        "cell_graph": all_cells[:max_cells],
        "sheets": sheets,
    }
    profile["signature"] = _build_signature({
        "sheets": [sheet.get("layout_signature", {}) for sheet in sheets],
        "cell_graph": profile["cell_graph"],
    })
    return profile


def compare_layout_profiles(base_profile, candidate_profile, compatible_threshold=0.78):
    if not isinstance(base_profile, dict) or not isinstance(candidate_profile, dict):
        return {
            "compatibility": "error",
            "score": 0.0,
            "reason": "Invalid profile payload",
            "hard_mismatch": True,
            "base_signature": None,
            "candidate_signature": None,
            "details": [],
        }

    base_signature = base_profile.get("signature")
    candidate_signature = candidate_profile.get("signature")
    if base_signature and candidate_signature and base_signature == candidate_signature:
        return {
            "compatibility": "exact",
            "score": 1.0,
            "reason": "Signature is identical",
            "hard_mismatch": False,
            "base_signature": base_signature,
            "candidate_signature": candidate_signature,
            "details": [],
        }

    base_map = _collect_sheet_map(base_profile)
    candidate_map = _collect_sheet_map(candidate_profile)
    sheet_names = set(base_map.keys()) | set(candidate_map.keys())
    details = []
    score_sum = 0.0

    for name in sorted(sheet_names):
        base_sheet = base_map.get(name)
        candidate_sheet = candidate_map.get(name)
        if base_sheet is None:
            details.append({"name": name, "reason": "missing_in_base", "score": 0.0})
            continue
        if candidate_sheet is None:
            details.append({"name": name, "reason": "missing_in_candidate", "score": 0.0})
            continue
        detail = _compare_sheet_sigs(base_sheet.get("layout_signature", {}), candidate_sheet.get("layout_signature", {}))
        details.append(detail)
        score_sum += detail["score"]

    score = score_sum / max(len(sheet_names), 1)
    compatibility = "compatible" if score >= compatible_threshold else "incompatible"
    return {
        "compatibility": compatibility,
        "score": round(score, 4),
        "reason": "Layout similarity below threshold" if compatibility == "incompatible" else "Layout similarity is acceptable",
        "hard_mismatch": compatibility == "incompatible",
        "base_signature": base_signature,
        "candidate_signature": candidate_signature,
        "details": details,
    }


def _load_profile(path):
    if path.suffix.lower() in {".xlsx", ".xls", ".xlsm", ".xlsb", ".xltx", ".xltm"}:
        return build_template_profile(str(path))
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--template', required=True)
    parser.add_argument('--output', required=True)
    parser.add_argument('--compare-with', required=False)
    parser.add_argument('--compatible-threshold', type=float, default=0.78)
    parser.add_argument('--max-cells', type=int, default=5000)
    args = parser.parse_args()

    profile = build_template_profile(args.template, max_cells=args.max_cells)
    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(profile, f, ensure_ascii=False, indent=2)

    if args.compare_with:
        other_profile = _load_profile(Path(args.compare_with))
        result = compare_layout_profiles(profile, other_profile, compatible_threshold=args.compatible_threshold)
        print(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()


def calculate_input_signature(file_path: str) -> str:
    """Wrapper around core signature calculation for input files."""
    from scripts.core.signature_matcher import calculate_signature
    with open(file_path, "r", encoding="utf-8") as f:
        return calculate_signature(f.read())
