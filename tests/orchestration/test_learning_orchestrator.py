# tests/orchestration/test_learning_orchestrator.py
import json
import tempfile
from pathlib import Path
import pytest
import openpyxl
from scripts.orchestration.learning_orchestrator import RALPHLoop

def test_ralph_loop_analyze_phase():
    """Test Phase 1: ANALYZE - LLM analyzes sample and generates anchors/schema."""
    with tempfile.TemporaryDirectory() as tmpdir:
        template_dir = Path(tmpdir)

        # Create sample input and target
        sample_input = template_dir / "sample.pdf"
        sample_input.write_text("Invoice #12345\nVendor: ABC Corp\nTotal: $1,000")

        target_excel = template_dir / "target.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B2"] = "12345"
        ws["B3"] = "ABC Corp"
        ws["B4"] = "1000"
        ws["D5"] = "=SUM(D2:D4)"
        wb.save(target_excel)

        ralph = RALPHLoop(str(template_dir))

        # Mock LLM response for analysis
        def mock_analyze(content, target_info):
            return {
                "anchors": {
                    "invoice_number": {
                        "type": "regex",
                        "pattern": "Invoice #(\\d+)",
                        "role": "value_matcher"
                    },
                    "vendor": {
                        "type": "text_match",
                        "patterns": ["Vendor:"],
                        "role": "label"
                    }
                },
                "fields": {
                    "Invoice_Number": {"cell": "B2", "type": "string"},
                    "Vendor_Name": {"cell": "B3", "type": "string"}
                },
                "formula_constraints": [
                    {"cell": "D5", "formula": "=SUM(D2:D4)"}
                ]
            }

        ralph._llm_analyze = mock_analyze

        result = ralph.phase1_analyze(str(sample_input), str(target_excel))

        assert "anchors" in result
        assert "fields" in result
        assert result["anchors"]["invoice_number"]["type"] == "regex"


def test_ralph_loop_draft_phase():
    """Test Phase 2: DRAFT - Generate initial rules from anchors."""
    with tempfile.TemporaryDirectory() as tmpdir:
        template_dir = Path(tmpdir)

        ralph = RALPHLoop(str(template_dir))

        anchors = {
            "invoice_number": {
                "type": "regex",
                "pattern": "Invoice #(\\d+)"
            }
        }
        fields = {
            "Invoice_Number": {"cell": "B2", "type": "string"}
        }

        def mock_draft_rules(anchors, fields):
            return [
                {
                    "id": "R001",
                    "when": {"input_type": "pdf", "trigger": "field_extraction"},
                    "condition": {"field": "Invoice_Number"},
                    "then": {"action": "extract_by_regex", "pattern": "Invoice #(\\d+)"},
                    "confidence": 0.9,
                    "support": 0
                }
            ]

        ralph._draft_rules = mock_draft_rules

        result = ralph.phase2_draft(anchors, fields)

        assert len(result) >= 1
        assert result[0]["id"] == "R001"
        assert result[0]["confidence"] >= 0.5


def test_ralph_loop_test_and_commit_phases():
    """Test Phase 3-4: TEST and COMMIT - Execute extraction and record success."""
    with tempfile.TemporaryDirectory() as tmpdir:
        template_dir = Path(tmpdir)

        # Setup template
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B2"] = ""
        ws["B3"] = ""
        wb.save(template_dir / "template.xlsx")

        schema = {
            "meta": {"signature": "test"},
            "fields": {
                "Invoice_Number": {"cell": "B2", "type": "string"},
                "Vendor_Name": {"cell": "B3", "type": "string"}
            }
        }
        with open(template_dir / "schema.json", "w") as f:
            json.dump(schema, f)

        ralph = RALPHLoop(str(template_dir))

        # Mock extraction that succeeds
        def mock_execute_extraction(content):
            return {"Invoice_Number": "12345", "Vendor_Name": "ABC Corp"}

        ralph._execute_extraction = mock_execute_extraction

        # Mock validation that passes
        def mock_validate(extracted, schema):
            return True, []

        ralph._validate = mock_validate

        test_input = "Invoice #12345\nVendor: ABC Corp"
        success, result = ralph.phase3_test(test_input, schema)

        assert success == True

        # Phase 4: COMMIT
        commit_result = ralph.phase4_commit(test_input, result)
        assert commit_result["status"] == "committed"


def test_ralph_loop_reflect_phase():
    """Test Phase 5: REFLECT - Analyze failure and update rules."""
    with tempfile.TemporaryDirectory() as tmpdir:
        template_dir = Path(tmpdir)

        ralph = RALPHLoop(str(template_dir))

        # Simulate a failure
        failure_info = {
            "missing_fields": ["Vendor_Name"],
            "error": "Extraction incomplete"
        }

        existing_rules = [
            {"id": "R001", "confidence": 0.9, "support": 5}
        ]

        def mock_analyze_failure(failure_info, rules):
            return [
                {
                    "id": "R002",
                    "when": {"trigger": "field_extraction"},
                    "condition": {"field": "Vendor_Name"},
                    "then": {"action": "extract_after_anchor", "anchor": "Vendor:"},
                    "confidence": 0.7,
                    "support": 0
                }
            ]

        ralph._analyze_failure = mock_analyze_failure

        new_rules = ralph.phase5_reflect(failure_info, existing_rules)

        assert len(new_rules) >= 1
        # The custom analyzer's repair rule (R002) is present, alongside
        # the penalized existing rule (R001) — the penalty now runs even
        # on the custom-analyzer path (Codex review fix).
        new_ids = {r["id"] for r in new_rules}
        assert "R002" in new_ids


def test_ralph_loop_full_cycle():
    """Test complete RALPH Loop with all 5 phases."""
    with tempfile.TemporaryDirectory() as tmpdir:
        template_dir = Path(tmpdir)

        # Setup
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B2"] = ""
        wb.save(template_dir / "template.xlsx")

        sample_input = template_dir / "sample.txt"
        sample_input.write_text("Invoice #12345")

        target_excel = template_dir / "target.xlsx"
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2["B2"] = "12345"
        wb2.save(target_excel)

        ralph = RALPHLoop(str(template_dir))

        # Mock all phases
        ralph._llm_analyze = lambda c, t: {
            "anchors": {},
            "fields": {"Invoice_Number": {"cell": "B2", "type": "string"}},
            "formula_constraints": []
        }
        ralph._draft_rules = lambda a, f: [
            {"id": "R001", "when": {}, "condition": {}, "then": {}, "confidence": 0.9, "support": 0}
        ]
        ralph._execute_extraction = lambda c: {"Invoice_Number": "12345"}
        ralph._validate = lambda e, s: (True, [])

        result = ralph.run_full_cycle(str(sample_input), str(target_excel), max_retries=3)

        assert result["status"] in ["committed", "reflected"]
        assert result["phase_reached"] >= 4  # Should reach at least COMMIT phase


# ---------------------------------------------------------------------------
# phase2_draft: auto-create uses_anchor edges (slice 2)
# ---------------------------------------------------------------------------

def test_phase2_draft_creates_uses_anchor_edge(tmp_path):
    """phase2_draft auto-creates uses_anchor edges for rules that
    reference an anchor in their field_spec.

    Given a field "Invoice_Number" whose field_spec carries anchor="A1",
    phase2_draft should:
    1. Build a rule whose `then` dict carries anchor="A1".
    2. Persist a {from: rule_id, to: "A1", relation: "uses_anchor"} edge
       in the knowledge graph.

    This is the write-side that makes slice 2's KG expansion observable:
    without edges being written, _expand_via_kg would have no neighbors.
    """
    ralph = RALPHLoop(str(tmp_path))
    ralph.input_type = "pdf"

    fields = {
        "Invoice_Number": {"cell": "B2", "type": "string", "anchor": "A1"},
    }
    rules = ralph.phase2_draft({}, fields)

    assert len(rules) == 1
    rule = rules[0]
    # The rule carries the anchor reference in its `then` block — this
    # is what ties the rule to the anchor in downstream logic.
    assert rule["then"]["anchor"] == "A1"

    # The knowledge graph now has exactly one uses_anchor edge from
    # this rule to A1.
    graph = ralph.bank.load_knowledge_graph()
    matching = [
        e for e in graph.get("edges", [])
        if e.get("relation") == "uses_anchor"
        and e.get("from") == rule["id"]
        and e.get("to") == "A1"
    ]
    assert len(matching) == 1


def test_phase2_draft_edge_creation_idempotent(tmp_path):
    """Re-adding the same (rule_id, anchor_id, uses_anchor) edge is a
    no-op — the knowledge graph never contains duplicate edges.

    Scenario: phase2_draft runs once and creates rule R001 with anchor
    A1, persisting edge (R001, A1, uses_anchor). We then force a second
    phase2_draft call to reuse rule ID R001 (simulating a rerun over
    the same rule) for the same anchor. The edge must NOT be duplicated.
    """
    ralph = RALPHLoop(str(tmp_path))
    ralph.input_type = "pdf"

    fields = {"Invoice_Number": {"cell": "B2", "anchor": "A1"}}
    rules1 = ralph.phase2_draft({}, fields)
    rule_id = rules1[0]["id"]
    assert rule_id == "R001"

    # Sanity: edge exists after first draft.
    def _count_edges(rule_id, anchor):
        g = ralph.bank.load_knowledge_graph()
        return sum(
            1 for e in g.get("edges", [])
            if e.get("from") == rule_id
            and e.get("to") == anchor
            and e.get("relation") == "uses_anchor"
        )
    assert _count_edges(rule_id, "A1") == 1

    # Force phase2_draft to reuse the same rule ID (R001) on the next
    # call. This simulates a rerun that reprocesses the same rule.
    original_next = ralph.bank._next_rule_id
    ralph.bank._next_rule_id = lambda: rule_id
    try:
        ralph.phase2_draft({}, fields)
    finally:
        ralph.bank._next_rule_id = original_next

    # Edge count is still 1 — idempotent.
    assert _count_edges(rule_id, "A1") == 1


# ---------------------------------------------------------------------------
# phase5_reflect: new rules keep creation confidence (slice 4 bugfix)
# ---------------------------------------------------------------------------

def test_phase5_reflect_new_rules_keep_creation_confidence(tmp_path):
    """phase5_reflect creates new repair rules at 0.6 for each missing
    field. Those NEW rules must keep 0.6 — they must NOT receive the
    failure penalty applied to existing rules.

    Bug: the old code ran update_rule_confidence(r, 0.2) over ALL rules
    in new_rules (existing + newly created), so a brand-new rule created
    at 0.6 was immediately penalized to 0.55. A rule just invented to
    fix a failure hasn't been tested yet — it shouldn't be punished for
    that same failure.
    """
    ralph = RALPHLoop(str(tmp_path))
    ralph.input_type = "pdf"
    failure_info = {"missing_fields": ["Vendor_Name"], "error": "incomplete"}
    result = ralph.phase5_reflect(failure_info, existing_rules=[])
    # One new rule created for the missing field.
    assert len(result) == 1
    # The new rule keeps its creation confidence (0.6), not penalized to 0.55.
    assert result[0]["confidence"] == 0.6


def test_phase5_reflect_existing_rules_penalized(tmp_path):
    """Complement to the new-rules-keep-0.6 test: rules that EXISTED
    before the reflection still receive the failure penalty (-0.05).

    Together the two tests pin down the distinction: existing rules are
    penalized for the failure; new repair rules are not. A fix that
    skipped penalizing entirely would pass S4-10 but fail here.
    """
    ralph = RALPHLoop(str(tmp_path))
    ralph.input_type = "pdf"
    existing = [
        {"id": "R001", "confidence": 0.9, "support": 5},
        {"id": "R002", "confidence": 0.8, "support": 3},
    ]
    failure_info = {"missing_fields": ["Vendor_Name"], "error": "incomplete"}
    result = ralph.phase5_reflect(failure_info, existing_rules=existing)

    by_id = {r["id"]: r for r in result}
    # Existing rules penalized: 0.9 → 0.85, 0.8 → 0.75.
    assert by_id["R001"]["confidence"] == 0.85
    assert by_id["R002"]["confidence"] == 0.75
    # New repair rule created at 0.6, unpenalized.
    new_ids = [rid for rid in by_id if rid not in {"R001", "R002"}]
    assert len(new_ids) == 1
    assert by_id[new_ids[0]]["confidence"] == 0.6


def test_phase5_reflect_custom_analyzer_still_penalizes_existing(tmp_path):
    """When _analyze_failure is set, the failure penalty must STILL run
    on existing rules — the custom hook generates repairs, it doesn't
    opt out of the slice-4 lifecycle policy.

    Codex PR #3 review (P2): the early return for _analyze_failure
    skipped apply_outcome(Outcome.FAILURE), so existing rules were
    neither penalized nor archived on failures handled by a custom
    analyzer. The penalty now runs before the custom analyzer, which
    receives the already-penalized rules; phase5_reflect returns the
    penalized existing rules PLUS the custom repairs.
    """
    ralph = RALPHLoop(str(tmp_path))
    ralph.input_type = "pdf"
    existing = [{"id": "R001", "confidence": 0.9, "support": 5}]
    received = {}

    def custom_analyzer(failure_info, rules):
        # Capture what we were handed — should be the penalized rules.
        received["confidences"] = [r["confidence"] for r in rules]
        # Return a single repair rule (new, not on disk).
        return [{"id": "R002", "confidence": 0.6, "support": 0,
                 "when": {}, "condition": {}, "then": {}}]

    ralph._analyze_failure = custom_analyzer
    result = ralph.phase5_reflect({"missing_fields": ["x"]}, existing)
    by_id = {r["id"]: r for r in result}

    # The custom analyzer received already-penalized rules (0.9 → 0.85).
    assert received["confidences"] == [0.85]
    # The existing rule is penalized AND still present (not dropped).
    assert "R001" in by_id
    assert by_id["R001"]["confidence"] == 0.85
    # The custom repair is included alongside the penalized existing rule.
    assert "R002" in by_id
    assert by_id["R002"]["confidence"] == 0.6


def test_phase3_test_threads_signature_into_retrieval(tmp_path, monkeypatch):
    """phase3_test passes schema.meta.signature into retrieve_rules so
    signature preference actually fires during the learning loop.

    Codex PR #3 review (P2): both orchestrators called retrieve_rules
    without input_signature, so via_signature promotion never happened
    in production. R001 here has input_type="pdf" — it can NEVER match
    an "md" query directly. It can only enter the retrieved set via
    signature preference (a matched pattern names it in rules_used).
    Observing that R001's last_used was freshened proves the signature
    was threaded: retrieve_rules only freshens last_used on rules it
    actually returns.
    """
    ralph = RALPHLoop(str(tmp_path))
    ralph.input_type = "md"
    # R001 is pdf — invisible to a direct "md" query.
    ralph.bank.save_rules([{
        "id": "R001",
        "when": {"input_type": "pdf", "trigger": "field_extraction"},
        "condition": {"field": "x"}, "then": {"action": "semantic_extract"},
        "confidence": 0.8, "support": 0,
    }])
    ralph.bank.save_success_patterns([
        {"pattern_id": "P001", "input_signature": "S1",
         "input_type": "pdf", "accuracy": 1.0, "rules_used": ["R001"]},
    ])
    schema = {"meta": {"signature": "S1"},
              "fields": {"x": {"cell": "A1", "type": "string"}}}

    # Drive phase3_test down its default (prompt-building) path, which
    # is where retrieve_rules is called. extract_data is mocked.
    import scripts.extraction.llm_extractor as llm
    monkeypatch.setattr(llm, "extract_data", lambda prompt: {"x": "v"})
    ralph._validate = lambda extracted, schema: (True, [])

    ralph.phase3_test("dummy content", schema)

    # R001 was retrieved via signature → its last_used was freshened.
    on_disk = ralph.bank.load_rules()
    assert on_disk[0].get("last_used") is not None


def test_phase4_commit_records_retrieved_rule_ids_in_success_pattern(tmp_path, monkeypatch):
    """phase4_commit records the rule IDs phase3_test retrieved into the
    success pattern's rules_used.

    Codex PR #3 review (P2): phase4_commit recorded rules_used=[] because
    phase3_test consumed the retrieved rules internally and never handed
    them back. With rules_used always empty, signature preference (slice
    3) could never accumulate rule associations in the learning loop —
    the via_signature path was dead in self-learning.

    R001 is a direct match for input_type="md". phase3_test retrieves it
    (and stashes the IDs); phase4_commit then writes a success pattern
    whose rules_used names R001.
    """
    ralph = RALPHLoop(str(tmp_path))
    ralph.input_type = "md"
    ralph.bank.save_rules([{
        "id": "R001",
        "when": {"input_type": "md", "trigger": "field_extraction"},
        "condition": {"field": "x"}, "then": {"action": "semantic_extract"},
        "confidence": 0.8, "support": 0,
    }])
    schema = {"meta": {"signature": "S1"},
              "fields": {"x": {"cell": "A1", "type": "string"}}}

    import scripts.extraction.llm_extractor as llm
    monkeypatch.setattr(llm, "extract_data", lambda prompt: {"x": "v"})
    ralph._validate = lambda extracted, schema: (True, [])

    success, extracted = ralph.phase3_test("dummy content", schema)
    assert success
    ralph.phase4_commit("dummy content", extracted)

    patterns = ralph.bank.load_success_patterns()
    assert len(patterns) == 1
    assert "R001" in patterns[0].get("rules_used", [])