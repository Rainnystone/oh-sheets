import json
import subprocess
import tempfile
import sys
import os
from pathlib import Path
import openpyxl

def _write_template(path: Path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Field_A"
    sheet["A2"] = "Field_B"
    workbook.save(path)

def _write_schema(path: Path):
    schema = {
        "meta": {
            "version": "2",
            "signature": "abc",
        },
        "fields": {
            "Field_A": {"cell": "B2", "type": "string"},
            "Field_B": {"relative_to": "Field_A", "row_offset": 1, "col_offset": 0, "type": "string"},
        },
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(schema, f)

def test_missing_field_fails_validation():
    with tempfile.TemporaryDirectory() as workdir:
        template_dir = Path(workdir)
        template_path = template_dir / "template.xlsx"
        schema_path = template_dir / "schema.json"
        input_path = template_dir / "input.dat"
        output_path = template_dir / "out.xlsx"

        _write_template(template_path)
        _write_schema(schema_path)

        input_path.write_text("dummy content")

        mock_extractor_path = template_dir / "mock_extractor.py"
        mock_extractor_content = """import sys, json
import scripts.orchestration.execution_orchestrator as eo

def mock_extract(prompt):
    return {"Field_A": "Only-first"}

eo.extract_data = mock_extract

if __name__ == "__main__":
    class Args:
        template_dir = sys.argv[1]
        input = sys.argv[2]
        output = sys.argv[3]
    sys.exit(eo.run_orchestrator(Args()))
"""
        mock_extractor_path.write_text(mock_extractor_content)

        result = subprocess.run(
            [sys.executable, str(mock_extractor_path), str(template_dir), str(input_path), str(output_path)],
            capture_output=True,
            text=True,
            env={"PYTHONPATH": ".", **os.environ}
        )

        assert result.returncode == 1
        payload = json.loads(result.stdout)
        assert payload["status"] == "failed"
        assert payload["stage"] == "validation"
        assert "Field_B" in payload["missing_fields"]


def test_successful_run_writes_output():
    with tempfile.TemporaryDirectory() as workdir:
        template_dir = Path(workdir)
        template_path = template_dir / "template.xlsx"
        schema_path = template_dir / "schema.json"
        input_path = template_dir / "input.dat"
        output_path = template_dir / "out.xlsx"

        _write_template(template_path)
        _write_schema(schema_path)
        input_path.write_text("dummy content")

        mock_extractor_path = template_dir / "mock_extractor.py"
        mock_extractor_content = """import sys, json
import scripts.orchestration.execution_orchestrator as eo
import scripts.io.excel_writer

def mock_extract(prompt):
    return {"Field_A": "A", "Field_B": "B"}

eo.extract_data = mock_extract
eo.write_excel = lambda t, d, s, o: open(o, "w").write("dummy excel")
scripts.io.excel_writer.write_excel = lambda t, d, s, o: open(o, "w").write("dummy excel")

if __name__ == "__main__":
    class Args:
        template_dir = sys.argv[1]
        input = sys.argv[2]
        output = sys.argv[3]
    sys.exit(eo.run_orchestrator(Args()))
"""
        mock_extractor_path.write_text(mock_extractor_content)

        result = subprocess.run(
            [sys.executable, str(mock_extractor_path), str(template_dir), str(input_path), str(output_path)],
            capture_output=True,
            text=True,
            env={"PYTHONPATH": ".", **os.environ}
        )

        assert result.returncode == 0
        payload = json.loads(result.stdout)
        assert payload["status"] == "success"


def test_memory_dir_uses_template_dir():
    """Test that execution log is written to template_dir/memory, not hardcoded 'memory'."""
    with tempfile.TemporaryDirectory() as workdir:
        template_dir = Path(workdir)
        template_path = template_dir / "template.xlsx"
        schema_path = template_dir / "schema.json"
        input_path = template_dir / "input.dat"
        output_path = template_dir / "out.xlsx"

        _write_template(template_path)
        _write_schema(schema_path)
        input_path.write_text("dummy content")

        mock_extractor_path = template_dir / "mock_extractor.py"
        mock_extractor_content = """import sys, json
import scripts.orchestration.execution_orchestrator as eo
import scripts.io.excel_writer

def mock_extract(prompt):
    return {"Field_A": "A", "Field_B": "B"}

eo.extract_data = mock_extract
scripts.io.excel_writer.write_excel = lambda t, d, s, o: open(o, "w").write("dummy excel")

if __name__ == "__main__":
    class Args:
        template_dir = sys.argv[1]
        input = sys.argv[2]
        output = sys.argv[3]
    sys.exit(eo.run_orchestrator(Args()))
"""
        mock_extractor_path.write_text(mock_extractor_content)

        result = subprocess.run(
            [sys.executable, str(mock_extractor_path), str(template_dir), str(input_path), str(output_path)],
            capture_output=True,
            text=True,
            env={"PYTHONPATH": ".", **os.environ}
        )

        # Execution log should be in template_dir/memory, not in cwd/memory
        expected_log_path = template_dir / "memory" / "execution_log.jsonl"
        assert expected_log_path.exists(), f"Expected log at {expected_log_path}"

        # Should NOT exist in cwd/memory
        cwd_memory_path = Path("memory") / "execution_log.jsonl"
        assert not cwd_memory_path.exists(), "Memory should not be written to cwd/memory"


def test_signature_mismatch_detection():
    """Test that new variant (no matching signature) is detected."""
    with tempfile.TemporaryDirectory() as workdir:
        template_dir = Path(workdir)
        template_path = template_dir / "template.xlsx"
        schema_path = template_dir / "schema.json"
        input_path = template_dir / "input.dat"
        output_path = template_dir / "out.xlsx"

        # Create reference_bank with existing patterns
        bank_dir = template_dir / "reference_bank"
        bank_dir.mkdir()
        import json
        patterns_path = bank_dir / "success_patterns.jsonl"
        # Existing pattern with different signature
        patterns_path.write_text(json.dumps({
            "input_signature": "known_signature_123",
            "accuracy": 0.95
        }) + "\n")

        _write_template(template_path)
        _write_schema(schema_path)
        # Input content will have a different signature
        input_path.write_text("completely different content for new signature")

        mock_extractor_path = template_dir / "mock_extractor.py"
        mock_extractor_content = """import sys, json
import scripts.orchestration.execution_orchestrator as eo
import scripts.io.excel_writer

def mock_extract(prompt):
    return {"Field_A": "A", "Field_B": "B"}

eo.extract_data = mock_extract
scripts.io.excel_writer.write_excel = lambda t, d, s, o: open(o, "w").write("dummy excel")

if __name__ == "__main__":
    class Args:
        template_dir = sys.argv[1]
        input = sys.argv[2]
        output = sys.argv[3]
    sys.exit(eo.run_orchestrator(Args()))
"""
        mock_extractor_path.write_text(mock_extractor_content)

        result = subprocess.run(
            [sys.executable, str(mock_extractor_path), str(template_dir), str(input_path), str(output_path)],
            capture_output=True,
            text=True,
            env={"PYTHONPATH": ".", **os.environ}
        )

        # Should succeed but indicate signature mismatch (new variant)
        payload = json.loads(result.stdout)
        assert payload["status"] == "success"
        assert "signature_mismatch" in payload or payload.get("variant") == "new"


def test_degradation_strategy_level2():
    """Test degradation to Level 2 (LLM + anchors only) when Level 1 fails."""
    with tempfile.TemporaryDirectory() as workdir:
        template_dir = Path(workdir)
        template_path = template_dir / "template.xlsx"
        schema_path = template_dir / "schema.json"
        input_path = template_dir / "input.dat"
        output_path = template_dir / "out.xlsx"

        _write_template(template_path)
        _write_schema(schema_path)
        input_path.write_text("test content")

        mock_extractor_path = template_dir / "mock_extractor.py"
        # Level 1 fails, Level 2 succeeds
        mock_extractor_content = """import sys, json
import scripts.orchestration.execution_orchestrator as eo
import scripts.io.excel_writer

call_count = [0]

def mock_extract(prompt):
    call_count[0] += 1
    # First call (Level 1 with rules) fails
    if call_count[0] == 1:
        raise Exception("Level 1 failed")
    # Second call (Level 2 anchors only) succeeds
    return {"Field_A": "A", "Field_B": "B"}

eo.extract_data = mock_extract
scripts.io.excel_writer.write_excel = lambda t, d, s, o: open(o, "w").write("dummy excel")

if __name__ == "__main__":
    class Args:
        template_dir = sys.argv[1]
        input = sys.argv[2]
        output = sys.argv[3]
    sys.exit(eo.run_orchestrator(Args()))
"""
        mock_extractor_path.write_text(mock_extractor_content)

        result = subprocess.run(
            [sys.executable, str(mock_extractor_path), str(template_dir), str(input_path), str(output_path)],
            capture_output=True,
            text=True,
            env={"PYTHONPATH": ".", **os.environ}
        )

        payload = json.loads(result.stdout)
        # Should succeed with degradation
        assert payload["status"] in ["success", "degraded"]
        assert payload.get("degraded_level") == 2 or "degraded" in payload.get("message", "")


def test_formula_conflict_detection():
    """Test that formula cells are detected and protected from being overwritten."""
    with tempfile.TemporaryDirectory() as workdir:
        template_dir = Path(workdir)
        template_path = template_dir / "template.xlsx"
        schema_path = template_dir / "schema.json"
        input_path = template_dir / "input.dat"
        output_path = template_dir / "out.xlsx"

        # Create template with formula
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["B2"] = ""  # Field_A cell
        ws["D5"] = "=SUM(D2:D4)"  # Formula cell - should not be overwritten
        wb.save(template_path)

        schema = {
            "meta": {"version": "2", "signature": "abc"},
            "fields": {
                "Field_A": {"cell": "B2", "type": "string"},
                "Field_B": {"cell": "D5", "type": "number"}  # This targets a formula cell
            }
        }
        with open(schema_path, "w") as f:
            json.dump(schema, f)

        input_path.write_text("test content")

        mock_extractor_path = template_dir / "mock_extractor.py"
        mock_extractor_content = """import sys, json
import scripts.orchestration.execution_orchestrator as eo
import scripts.io.excel_writer

def mock_extract(prompt):
    # Returns data that includes Field_B which targets a formula cell
    return {"Field_A": "Test Value", "Field_B": 100}

eo.extract_data = mock_extract
scripts.io.excel_writer.write_excel = lambda t, d, s, o: open(o, "w").write("dummy excel")

if __name__ == "__main__":
    class Args:
        template_dir = sys.argv[1]
        input = sys.argv[2]
        output = sys.argv[3]
    sys.exit(eo.run_orchestrator(Args()))
"""
        mock_extractor_path.write_text(mock_extractor_content)

        result = subprocess.run(
            [sys.executable, str(mock_extractor_path), str(template_dir), str(input_path), str(output_path)],
            capture_output=True,
            text=True,
            env={"PYTHONPATH": ".", **os.environ}
        )

        payload = json.loads(result.stdout)
        # Should warn about formula conflict
        assert payload["status"] in ["success", "degraded"]
        # Check that formula conflict is detected
        assert payload.get("formula_conflict") == True or "formula" in payload.get("message", "").lower()


def test_run_orchestrator_threads_signature_into_retrieve_rules():
    """run_orchestrator passes the input content's signature into
    retrieve_rules so via_signature promotion fires in production.

    Codex PR #3 review (P2): the execution orchestrator called
    retrieve_rules(input_type) without input_signature, so rules
    reachable only via a matched success pattern were never retrieved.
    R001 here is input_type="pdf" — invisible to an "md" (.dat) query
    directly. A success pattern whose input_signature equals the MD5 of
    the input content names R001 in rules_used. Observing that R001's
    last_used was freshened proves the signature was threaded:
    retrieve_rules only freshens last_used on rules it returns.
    """
    import hashlib
    with tempfile.TemporaryDirectory() as workdir:
        template_dir = Path(workdir)
        template_path = template_dir / "template.xlsx"
        schema_path = template_dir / "schema.json"
        input_path = template_dir / "input.dat"
        output_path = template_dir / "out.xlsx"

        _write_template(template_path)
        _write_schema(schema_path)

        content = "dummy content for signature threading"
        input_path.write_text(content)
        input_sig = hashlib.md5(content.encode("utf-8")).hexdigest()

        # Pre-populate the bank: R001 is pdf (invisible to .dat -> "md").
        bank_dir = template_dir / "reference_bank"
        bank_dir.mkdir()
        (bank_dir / "rules.jsonl").write_text(json.dumps({
            "id": "R001",
            "when": {"input_type": "pdf", "trigger": "field_extraction"},
            "condition": {"field": "Field_A"},
            "then": {"action": "semantic_extract"},
            "confidence": 0.8,
            "support": 0,
        }) + "\n")
        # Pattern whose signature matches the input, naming R001.
        (bank_dir / "success_patterns.jsonl").write_text(json.dumps({
            "pattern_id": "P001",
            "input_signature": input_sig,
            "input_type": "pdf",
            "accuracy": 1.0,
            "rules_used": ["R001"],
        }) + "\n")

        mock_extractor_path = template_dir / "mock_extractor.py"
        mock_extractor_content = """import sys, json
import scripts.orchestration.execution_orchestrator as eo
import scripts.io.excel_writer

def mock_extract(prompt):
    return {"Field_A": "A", "Field_B": "B"}

eo.extract_data = mock_extract
scripts.io.excel_writer.write_excel = lambda t, d, s, o: open(o, "w").write("dummy excel")

if __name__ == "__main__":
    class Args:
        template_dir = sys.argv[1]
        input = sys.argv[2]
        output = sys.argv[3]
    sys.exit(eo.run_orchestrator(Args()))
"""
        mock_extractor_path.write_text(mock_extractor_content)

        result = subprocess.run(
            [sys.executable, str(mock_extractor_path), str(template_dir), str(input_path), str(output_path)],
            capture_output=True,
            text=True,
            env={"PYTHONPATH": ".", **os.environ}
        )

        # R001 was retrieved via signature -> its last_used was freshened on disk.
        from scripts.core.reference_bank import ReferenceBank
        bank = ReferenceBank(str(bank_dir))
        rules = bank.load_rules()
        assert rules[0].get("last_used") is not None, (
            f"R001 last_used not freshened — signature not threaded. "
            f"stdout={result.stdout!r} stderr={result.stderr!r}"
        )


def test_degraded_extraction_omits_rules_used():
    """A degraded (level>1) extraction records rules_used=[] because the
    rules were never sent to the LLM.

    Codex PR #3 review (P2): run_orchestrator recorded the full retrieved
    set in rules_used even when degradation dropped to Level 2 (LLM +
    anchors only, rules=[] in the prompt). That corrupted signature
    preference — a future input matching this signature would be told
    "these rules worked" when they were never in the prompt.

    R001 is retrieved at the top (input_type="md" matches .dat) and would
    be in the Level 1 prompt, but Level 1 fails and the run degrades to
    Level 2 (no rules). The recorded success pattern must therefore have
    rules_used=[].
    """
    with tempfile.TemporaryDirectory() as workdir:
        template_dir = Path(workdir)
        template_path = template_dir / "template.xlsx"
        schema_path = template_dir / "schema.json"
        input_path = template_dir / "input.dat"
        output_path = template_dir / "out.xlsx"

        _write_template(template_path)
        _write_schema(schema_path)
        input_path.write_text("test content")

        # Pre-populate the bank: R001 matches .dat -> "md" so it's retrieved.
        bank_dir = template_dir / "reference_bank"
        bank_dir.mkdir()
        (bank_dir / "rules.jsonl").write_text(json.dumps({
            "id": "R001",
            "when": {"input_type": "md", "trigger": "field_extraction"},
            "condition": {"field": "Field_A"},
            "then": {"action": "semantic_extract"},
            "confidence": 0.8,
            "support": 0,
        }) + "\n")

        mock_extractor_path = template_dir / "mock_extractor.py"
        # Level 1 (with rules) fails, Level 2 (anchors only) succeeds.
        mock_extractor_content = """import sys, json
import scripts.orchestration.execution_orchestrator as eo
import scripts.io.excel_writer

call_count = [0]

def mock_extract(prompt):
    call_count[0] += 1
    if call_count[0] == 1:
        raise Exception("Level 1 failed")
    return {"Field_A": "A", "Field_B": "B"}

eo.extract_data = mock_extract
scripts.io.excel_writer.write_excel = lambda t, d, s, o: open(o, "w").write("dummy excel")

if __name__ == "__main__":
    class Args:
        template_dir = sys.argv[1]
        input = sys.argv[2]
        output = sys.argv[3]
    sys.exit(eo.run_orchestrator(Args()))
"""
        mock_extractor_path.write_text(mock_extractor_content)

        result = subprocess.run(
            [sys.executable, str(mock_extractor_path), str(template_dir), str(input_path), str(output_path)],
            capture_output=True,
            text=True,
            env={"PYTHONPATH": ".", **os.environ}
        )

        payload = json.loads(result.stdout)
        # Degraded to level 2 — succeeded without rules in the prompt.
        assert payload.get("degraded_level") == 2

        # The recorded success pattern must NOT name R001 — it was never
        # sent to the LLM (Level 2 uses rules=[]).
        from scripts.core.reference_bank import ReferenceBank
        bank = ReferenceBank(str(bank_dir))
        patterns = bank.load_success_patterns()
        assert len(patterns) == 1
        assert patterns[0].get("rules_used") == [], (
            f"degraded extraction recorded rules_used="
            f"{patterns[0].get('rules_used')!r} — should be empty"
        )


class _StubBank:
    """Duck-typed stand-in for ReferenceBank.

    _retrieve_context accepts the bank as a dependency (instead of
    constructing one internally), so a unit test can pass this stub and
    observe exactly which methods were called with which arguments —
    no on-disk ReferenceBank required.
    """
    def __init__(self, rules=None, anchors=None, patterns=None):
        self._rules = rules if rules is not None else []
        self._anchors = anchors if anchors is not None else {}
        self._patterns = patterns if patterns is not None else []
        self.received_input_type = None
        self.received_input_signature = None

    def retrieve_rules(self, input_type, input_signature=None):
        self.received_input_type = input_type
        self.received_input_signature = input_signature
        return self._rules

    def load_anchors(self):
        return self._anchors

    def load_success_patterns(self):
        return self._patterns


def test_retrieve_context_assembles_context_when_signature_known():
    """_retrieve_context builds the full extraction context — signature,
    input_type, rules, anchors, patterns, matched, signature_mismatch —
    behind one call, threading (input_type, input_signature) into
    retrieve_rules.

    Slice 6b (god-function decomposition, candidate 02). With a known
    signature stored in the bank, the input is NOT a new variant.
    """
    import hashlib
    from scripts.orchestration.execution_orchestrator import _retrieve_context

    content = "hello world"
    input_sig = hashlib.md5(content.encode("utf-8")).hexdigest()
    rules = [{"id": "R001"}]
    anchors = {"anchor_a": {"row": 2}}
    patterns = [{"input_signature": input_sig, "accuracy": 1.0}]

    bank = _StubBank(rules=rules, anchors=anchors, patterns=patterns)
    ctx = _retrieve_context(bank, content, "sample.pdf")

    assert ctx.input_sig == input_sig
    assert ctx.input_type == "pdf"
    assert ctx.rules == rules
    assert ctx.anchors == anchors
    assert ctx.patterns == patterns
    # exact signature match -> the stored pattern is returned by match_patterns
    assert ctx.matched == patterns
    assert ctx.signature_mismatch is False
    # the bank was queried with the right (input_type, input_signature)
    assert bank.received_input_type == "pdf"
    assert bank.received_input_signature == input_sig


def test_retrieve_context_flags_new_variant_when_signature_unseen():
    """When stored patterns exist but none match the input signature,
    signature_mismatch is True (new variant detected)."""
    from scripts.orchestration.execution_orchestrator import _retrieve_context

    patterns = [{"input_signature": "some_other_signature", "accuracy": 0.9}]
    bank = _StubBank(patterns=patterns)

    ctx = _retrieve_context(bank, "a brand new variant", "sample.md")

    assert ctx.input_type == "md"
    assert ctx.signature_mismatch is True
    # no exact match -> match_patterns falls back to accuracy >= threshold
    assert ctx.matched == patterns


def test_retrieve_context_no_mismatch_when_no_patterns_stored():
    """With no stored patterns there is nothing to compare against, so
    signature_mismatch stays False (the `patterns and ...` guard)."""
    from scripts.orchestration.execution_orchestrator import _retrieve_context

    bank = _StubBank(patterns=[])
    ctx = _retrieve_context(bank, "fresh start, no history", "sample.txt")

    assert ctx.signature_mismatch is False
    assert ctx.matched == []
    assert ctx.input_type == "md"


def test_detect_formula_conflicts_removes_field_mapping_to_formula_cell(tmp_path):
    """A schema field whose target cell holds a formula is reported as a
    conflict and dropped from the extracted data so the formula cell is
    never overwritten.

    Slice 6c (god-function decomposition, candidate 02).
    """
    import openpyxl
    from scripts.orchestration.execution_orchestrator import _detect_formula_conflicts

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["D5"] = "=SUM(D2:D4)"  # formula cell
    wb.save(str(tmp_path / "template.xlsx"))

    schema = {
        "fields": {
            "Field_A": {"cell": "B2", "type": "string"},   # plain cell
            "Field_B": {"cell": "D5", "type": "number"},    # formula cell -> conflict
        }
    }
    extracted = {"Field_A": "keep me", "Field_B": 100}

    result_extracted, conflicts = _detect_formula_conflicts(schema, extracted, tmp_path)

    assert conflicts == [{"field": "Field_B", "cell": "D5"}]
    assert "Field_B" not in result_extracted
    assert result_extracted["Field_A"] == "keep me"


def test_detect_formula_conflicts_no_conflict_when_cell_is_plain(tmp_path):
    """No conflict when schema fields map to plain (non-formula) cells."""
    import openpyxl
    from scripts.orchestration.execution_orchestrator import _detect_formula_conflicts

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B2"] = ""
    wb.save(str(tmp_path / "template.xlsx"))

    schema = {"fields": {"Field_A": {"cell": "B2", "type": "string"}}}
    extracted = {"Field_A": "value"}

    result_extracted, conflicts = _detect_formula_conflicts(schema, extracted, tmp_path)

    assert conflicts == []
    assert result_extracted == {"Field_A": "value"}


def test_detect_formula_conflicts_swallows_analysis_failure(tmp_path):
    """If formula analysis fails (e.g. no template.xlsx present), the step
    returns the extracted data untouched with no conflicts — it never
    blocks the run. Mirrors the original `except Exception: pass` guard.
    """
    from scripts.orchestration.execution_orchestrator import _detect_formula_conflicts

    schema = {"fields": {"Field_A": {"cell": "B2", "type": "string"}}}
    extracted = {"Field_A": "value"}
    # tmp_path has no template.xlsx -> analyze_workbook_formulas raises

    result_extracted, conflicts = _detect_formula_conflicts(schema, extracted, tmp_path)

    assert conflicts == []
    assert result_extracted == {"Field_A": "value"}


class _RecordingBank:
    """Duck-typed stand-in for ReferenceBank's success-side writes.

    Captures apply_outcome and record_success_pattern calls so a unit
    test can assert exactly what the bank was told — no on-disk
    ReferenceBank required.
    """
    def __init__(self):
        self.outcomes = []
        self.recorded_patterns = []

    def apply_outcome(self, outcome):
        self.outcomes.append(outcome)

    def record_success_pattern(self, **kwargs):
        self.recorded_patterns.append(kwargs)


def test_validate_returns_missing_fields():
    """_validate returns the list of required schema fields absent from
    the extracted data.

    Slice 6d (god-function decomposition, candidate 02).
    """
    from scripts.orchestration.execution_orchestrator import _validate

    schema = {"fields": {"Field_A": {"cell": "B2"}, "Field_B": {"cell": "C2"}}}
    extracted = {"Field_A": "v"}  # Field_B missing

    assert _validate(extracted, schema) == ["Field_B"]


def test_validate_returns_empty_when_all_fields_present():
    """When every required schema field is present, _validate returns []."""
    from scripts.orchestration.execution_orchestrator import _validate

    schema = {"fields": {"Field_A": {"cell": "B2"}, "Field_B": {"cell": "C2"}}}
    extracted = {"Field_A": "v", "Field_B": "w"}

    assert _validate(extracted, schema) == []


def test_record_outcome_rewards_rules_and_records_pattern():
    """_record_outcome applies SUCCESS to every rule and records a success
    pattern carrying the retrieved rule ids in rules_used (not degraded).

    Slice 6d (god-function decomposition, candidate 02).
    """
    from scripts.orchestration.execution_orchestrator import _record_outcome, _RetrievalContext
    from scripts.core.rule_evolution import Outcome

    bank = _RecordingBank()
    ctx = _RetrievalContext(
        rules=[{"id": "R001"}, {"id": "R002"}],
        anchors={"anchor_a": {"row": 1}},
        patterns=[],
        matched=[],
        input_sig="abc123",
        input_type="md",
        signature_mismatch=False,
    )

    _record_outcome(bank, ctx, extracted={"Field_A": "v"}, degraded_level=1)

    assert bank.outcomes == [Outcome.SUCCESS]
    assert len(bank.recorded_patterns) == 1
    pat = bank.recorded_patterns[0]
    assert pat["input_signature"] == "abc123"
    assert pat["input_type"] == "md"
    assert pat["extracted"] == {"Field_A": "v"}
    assert pat["rules_used"] == ["R001", "R002"]  # not degraded -> rules recorded
    assert pat["anchors_matched"] == ["anchor_a"]
    assert pat["accuracy"] == 1.0


def test_record_outcome_omits_rules_used_when_degraded():
    """A degraded (level>1) extraction never sent rules to the LLM, so
    rules_used must be [] to avoid corrupting signature preference
    (Codex PR #3 review P2)."""
    from scripts.orchestration.execution_orchestrator import _record_outcome, _RetrievalContext

    bank = _RecordingBank()
    ctx = _RetrievalContext(
        rules=[{"id": "R001"}],
        anchors={},
        patterns=[],
        matched=[],
        input_sig="abc123",
        input_type="md",
        signature_mismatch=False,
    )

    _record_outcome(bank, ctx, extracted={"Field_A": "v"}, degraded_level=2)

    pat = bank.recorded_patterns[0]
    assert pat["rules_used"] == []  # degraded -> rules never reached the prompt
