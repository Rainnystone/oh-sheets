import os
import json
import openpyxl
import subprocess

def test_excel_writer():
    # Setup blank template
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Header'
    wb.save('test_template.xlsx')
    
    # Setup JSON data and Schema
    data = {"A2": "Value1"}
    with open('test_data.json', 'w') as f:
        json.dump(data, f)
        
    schema = {"A2": {"name": "Field1", "type": "string"}}
    with open('test_schema.json', 'w') as f:
        json.dump(schema, f)
        
    result = subprocess.run(['python', 'scripts/io/excel_writer.py', '--template', 'test_template.xlsx', '--data', 'test_data.json', '--schema', 'test_schema.json', '--output', 'test_out.xlsx'])
    
    assert result.returncode == 0
    
    wb_out = openpyxl.load_workbook('test_out.xlsx')
    assert wb_out.active['A2'].value == "Value1"
    
    os.remove('test_template.xlsx')
    os.remove('test_data.json')
    os.remove('test_schema.json')
    os.remove('test_out.xlsx')


def test_excel_writer_with_semantic_layout():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Header'
    wb.save('test_template.xlsx')

    data = {
        "Field_A": "Value_A",
        "Field_B": "2026-01-01",
        "Field_C": "8.0",
    }

    schema = {
        "meta": {"version": "2", "signature": "signature"},
        "fields": {
            "Field_A": {"cell": "B2", "type": "string"},
            "Field_B": {"relative_to": "Field_A", "row_offset": 1, "col_offset": 0},
            "Field_C": {"relative_to": "Field_B", "row_offset": 0, "col_offset": 1},
        },
    }

    with open('test_data.json', 'w') as f:
        json.dump(data, f)
    with open('test_schema.json', 'w') as f:
        json.dump(schema, f)

    result = subprocess.run(['python', 'scripts/io/excel_writer.py', '--template', 'test_template.xlsx', '--data', 'test_data.json', '--schema', 'test_schema.json', '--output', 'test_out.xlsx'])
    assert result.returncode == 0

    wb_out = openpyxl.load_workbook('test_out.xlsx')
    assert wb_out.active['B2'].value == "Value_A"
    assert wb_out.active['B3'].value == "2026-01-01"
    assert wb_out.active['C3'].value == "8.0"

    os.remove('test_template.xlsx')
    os.remove('test_data.json')
    os.remove('test_schema.json')
    os.remove('test_out.xlsx')


def test_write_excel_returns_none_on_success(tmp_path):
    """write_excel is a library function: on success it returns None,
    not sys.exit(0). Callers (and tests) can invoke it directly without
    a SystemExit escaping the process.

    Candidate 04 (roadmap slice 5): sys.exit belongs only in __main__.
    """
    wb = openpyxl.Workbook()
    wb.active['A1'] = 'Header'
    template = tmp_path / "template.xlsx"
    wb.save(template)

    data = {"Field1": "Value1"}
    data_path = tmp_path / "data.json"
    data_path.write_text(json.dumps(data))

    schema = {"A2": {"name": "Field1", "type": "string"}}
    schema_path = tmp_path / "schema.json"
    schema_path.write_text(json.dumps(schema))

    output = tmp_path / "out.xlsx"

    from scripts.io.excel_writer import write_excel
    result = write_excel(str(template), str(data_path), str(schema_path), str(output))

    assert result is None
    assert output.exists()
    wb_out = openpyxl.load_workbook(output)
    assert wb_out.active['A2'].value == "Value1"


def test_write_excel_raises_on_failure_not_sys_exit(tmp_path):
    """On failure write_excel raises a normal exception, not SystemExit.

    Candidate 04 (roadmap slice 5): the old code caught every exception
    and called sys.exit(1), forcing callers to catch SystemExit. Now the
    exception propagates and only __main__ exits.
    """
    import pytest
    data_path = tmp_path / "data.json"
    data_path.write_text(json.dumps({"Field1": "v"}))
    schema_path = tmp_path / "schema.json"
    schema_path.write_text(json.dumps({"A2": {"name": "Field1"}}))

    from scripts.io.excel_writer import write_excel
    # Missing template file -> openpyxl raises (e.g. FileNotFoundError or
    # BadZipFile). Must NOT be SystemExit.
    with pytest.raises(Exception) as exc_info:
        write_excel(str(tmp_path / "nonexistent.xlsx"),
                    str(data_path), str(schema_path),
                    str(tmp_path / "out.xlsx"))
    assert not isinstance(exc_info.value, SystemExit)
