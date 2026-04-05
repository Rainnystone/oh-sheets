# tests/extraction/test_formula_analyzer.py
import os
import tempfile
from openpyxl import Workbook
from scripts.extraction.formula_analyzer import (
    extract_formulas_from_schema,
    analyze_workbook_formulas,
    extract_formula_dependencies
)

def test_extract_formulas_from_schema():
    """Test extracting formula constraints from schema definition."""
    schema = {
        "formula_constraints": [
            {"cell": "D5", "formula": "=SUM(D2:D4)"}
        ]
    }
    formulas = extract_formulas_from_schema(schema)
    assert len(formulas) == 1
    assert formulas[0]["cell"] == "D5"

def test_analyze_workbook_formulas():
    """Test extracting formulas from an actual Excel workbook."""
    # Create a test workbook with formulas
    wb = Workbook()
    ws = wb.active
    ws['D2'] = 100
    ws['D3'] = 200
    ws['D4'] = 300
    ws['D5'] = '=SUM(D2:D4)'  # Formula cell

    # Save to temp file
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp_path = tmp.name
        wb.save(tmp_path)

    try:
        formulas = analyze_workbook_formulas(tmp_path)

        # Should find the SUM formula
        assert len(formulas) >= 1
        formula_found = any(f['cell'] == 'D5' for f in formulas)
        assert formula_found, f"Expected formula at D5, got: {formulas}"

        # Formula should have correct structure
        d5_formula = next(f for f in formulas if f['cell'] == 'D5')
        assert 'formula' in d5_formula
        assert 'SUM' in d5_formula['formula'] or 'sum' in d5_formula['formula'].lower()
        assert 'depends_on' in d5_formula
    finally:
        os.unlink(tmp_path)

def test_extract_formula_dependencies():
    """Test extracting cell references from formula string."""
    deps = extract_formula_dependencies('=SUM(D2:D4)+A1')
    assert 'A1' in deps
    # Range references may include multiple cells
    assert len(deps) >= 1

def test_formula_cell_identification():
    """Test that formula cells are properly identified."""
    wb = Workbook()
    ws = wb.active
    ws['D5'] = '=SUM(D2:D4)'
    ws['A1'] = 'Static value'  # Not a formula

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp_path = tmp.name
        wb.save(tmp_path)

    try:
        formulas = analyze_workbook_formulas(tmp_path)
        formula_cells = [f['cell'] for f in formulas]

        # D5 should be identified as a formula cell
        assert 'D5' in formula_cells
        # A1 should NOT be in formula cells
        assert 'A1' not in formula_cells
    finally:
        os.unlink(tmp_path)
