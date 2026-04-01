import json
import os
import openpyxl
import subprocess


def _build_template_for_test(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TemplateSheet"
    ws['A1'] = 'Field_A'
    ws['B1'] = 'Field_B'
    ws['A2'] = 'v1'
    ws['B2'] = 8
    wb.save(path)


def test_template_layout_signature():
    template = 'test_layout_template.xlsx'
    profile = 'test_layout_profile.json'
    _build_template_for_test(template)

    result = subprocess.run(
        ['python', 'scripts/template_layout_signature.py', '--template', template, '--output', profile],
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0

    with open(profile, 'r', encoding='utf-8') as f:
        data = json.load(f)

    assert data['schema_version'] == '2.0'
    assert data['sheets'][0]['name'] == 'TemplateSheet'
    assert data['signature']
    assert any(cell['cell'] == 'A1' for cell in data['cell_graph'])

    result_repeat = subprocess.run(
        ['python', 'scripts/template_layout_signature.py', '--template', template, '--output', profile],
        capture_output=True,
        text=True,
    )
    assert result_repeat.returncode == 0

    with open(profile, 'r', encoding='utf-8') as f:
        data_repeat = json.load(f)

    assert data['signature'] == data_repeat['signature']

    os.remove(template)
    os.remove(profile)
