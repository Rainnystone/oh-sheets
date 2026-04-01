import os
import json
import openpyxl
import subprocess

def test_excel_writer():
    # Setup blank template
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Header'
    wb.save('test_template.xlsx')
    
    # Setup JSON data and Schema
    data = {"A2": "Value1"}
    with open('test_data.json', 'w') as f:
        json.dump(data, f)
        
    schema = {"A2": {"name": "Field1", "type": "string"}}
    with open('test_schema.json', 'w') as f:
        json.dump(schema, f)
        
    result = subprocess.run(['python', 'scripts/excel_writer.py', '--template', 'test_template.xlsx', '--data', 'test_data.json', '--schema', 'test_schema.json', '--output', 'test_out.xlsx'])
    
    assert result.returncode == 0
    
    wb_out = openpyxl.load_workbook('test_out.xlsx')
    assert wb_out.active['A2'].value == "Value1"
    
    os.remove('test_template.xlsx')
    os.remove('test_data.json')
    os.remove('test_schema.json')
    os.remove('test_out.xlsx')


def test_excel_writer_with_semantic_layout():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Header'
    wb.save('test_template.xlsx')

    data = {
        "Field_A": "Value_A",
        "Field_B": "2026-01-01",
        "Field_C": "8.0",
    }

    schema = {
        "meta": {"version": "2", "signature": "signature"},
        "fields": {
            "Field_A": {"cell": "B2", "type": "string"},
            "Field_B": {"relative_to": "Field_A", "row_offset": 1, "col_offset": 0},
            "Field_C": {"relative_to": "Field_B", "row_offset": 0, "col_offset": 1},
        },
    }

    with open('test_data.json', 'w') as f:
        json.dump(data, f)
    with open('test_schema.json', 'w') as f:
        json.dump(schema, f)

    result = subprocess.run(['python', 'scripts/excel_writer.py', '--template', 'test_template.xlsx', '--data', 'test_data.json', '--schema', 'test_schema.json', '--output', 'test_out.xlsx'])
    assert result.returncode == 0

    wb_out = openpyxl.load_workbook('test_out.xlsx')
    assert wb_out.active['B2'].value == "Value_A"
    assert wb_out.active['B3'].value == "2026-01-01"
    assert wb_out.active['C3'].value == "8.0"

    os.remove('test_template.xlsx')
    os.remove('test_data.json')
    os.remove('test_schema.json')
    os.remove('test_out.xlsx')
